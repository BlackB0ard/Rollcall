import cognitive_face as CF
import global_variables as global_var
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import cv2
import numpy as np
import pandas as pd
from datetime import datetime,timedelta
import shutil
from datetime import date,time
import dlib
import sys


requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

current_day=date.today().weekday()


now=datetime.now()
currentdate=now.strftime("%d_%m_%y")
conn = sqlite3.connect('Face-DataBase')
c = conn.cursor()


data=pd.read_csv("time_table (copy).csv")
currenttime=now.strftime("%H:%M")

dict={}
a=np.array(data.values[0][1:])
k=0

periods=data.values[current_day+1][1:]
for i in periods:
    dict[a[k]]=i
    k=k+1


def is_time_between(begin_time, end_time, check_time=None):
    # If check time is not given, default to current UTC time
    check_time = check_time or datetime.now().time()
    #print(check_time)
    if begin_time < end_time:
        return check_time >= begin_time and check_time <= end_time
    else: # crosses midnight
        return check_time >= begin_time or check_time <= end_time



def path12():
    if(is_time_between(time(20,0), time(20,9))):
        path="./"+currentdate+"/"+dict['20:00-20:09']+"/"+dict['20:00-20:09']+"_"+currentdate+".xlsx"

    if(is_time_between(time(20,10), time(20,20))):
        path="./"+currentdate+"/"+dict['20:10-20:20']+"/"+dict['20:10-20:19']+"_"+currentdate+".xlsx"


    if(is_time_between(time(20,20), time(20,29))):
        path="./"+currentdate+"/"+dict['20:20-20:29']+"/"+dict['20:20-20:29']+"_"+currentdate+".xlsx"


    if(is_time_between(time(20,30), time(20,39))):
        path="./"+currentdate+"/"+dict['20:30-20:39']+"/"+dict['20:30-20:39']+"_"+currentdate+".xlsx"


    if(is_time_between(time(20,40), time(20,49))):
        path="./"+currentdate+"/"+dict['20:40-20:49']+"/"+dict['20:40-20:49']+"_"+currentdate+".xlsx"


    if(is_time_between(time(20,50), time(20,59))):
        path="./"+currentdate+"/"+dict['20:50-20:59']+"/"+dict['20:50-20:59']+"_"+currentdate+".xlsx"

    return path

def identify(path,column):
    #get current date
    now=datetime.now()
    currentDate = now.strftime("%d_%m_%y")
    wb = load_workbook(path)
    sheet = wb['Cse16']

    def getColumn():
        for i in range(1, len(list(sheet.rows)[0]) + 1):
            col = get_column_letter(i)
            if sheet['%s%s'% (col,'1')].value == column:
                return col

    Key = global_var.key

    CF.Key.set(Key)

    BASE_URL = global_var.BASE_URL  # Replace with your regional Base URL
    CF.BaseUrl.set(BASE_URL)

    connect = sqlite3.connect("Face-DataBase")
    #c = connect.cursor()

    attend = [0 for i in range(60)]

    currentDir = os.path.dirname(os.path.abspath(''))+"/"+"final_project"
    directory = os.path.join(currentDir, 'Cropped_faces')
    for filename in os.listdir(directory):
        if filename.endswith(".jpg"):
            imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
            #imgurl = imgurl[3:]
            print("imgurl = {}".format(imgurl))
            res = CF.face.detect(imgurl)
            print("Res = {}".format(res))

            if len(res) < 1:
                print("No face detected.")
                continue

            faceIds = []
            for face in res:
                faceIds.append(face['faceId'])
            res = CF.face.identify(faceIds, global_var.personGroupId)
            print(filename)
            print("res = {}".format(res))

            for face  in res:
                if not face['candidates']:
                    print("Unknown")
                else:
                    personId = face['candidates'][0]['personId']
                    print("personid = {}".format(personId))
                    #cmd =  + personId
                    cur = connect.execute("SELECT * FROM Students WHERE personID = (?)", (personId,))
                    #print("cur = {}".format(cur))
                    for row in cur:
                        print("aya")
                        print("row = {}".format(row))
                        attend[int(row[0])] += 1
                    print("---------- " + row[1] + " recognized ----------")
            time.sleep(6)

    for row in range(2, len(list(sheet.columns)[0]) + 1):
        rn = sheet.cell(row = row, column  =1).value
        if rn is not None:
            print("rn = {}".format(rn))
            rn = rn[-2:]
            if attend[int(rn)] != 0:
                col = getColumn()
                print("col = {}".format(col))
                sheet['%s%s' % (col, str(row))] = 1

    wb.save(path)



def detect():

    detector = dlib.get_frontal_face_detector()

    cam = cv2.VideoCapture(0)
    detector = dlib.get_frontal_face_detector()
    ret, frame = cam.read()
    ret, frame = cam.read()
    ret, frame = cam.read()
    count=1
    cv2.imwrite("./pics/framee%d.jpg" % count, frame)

    img = cv2.imread('./pics/framee1.jpg')
    dets = detector(img, 1)
    if not os.path.exists('./Cropped_faces'):
        os.makedirs('./Cropped_faces')
        print("detected = {}".format(len(dets)))
        for i, d in enumerate(dets):
            cv2.imwrite('./Cropped_faces/face' + str(i + 1) + '.jpg', img[d.top():d.bottom(), d.left():d.right()])


while is_time_between(time(20,0), time(21,0)):
    hello12=datetime.now().time().strftime("%H:%M")
    #print(hello[3:5])
    path=path12()

    if(hello12[4:5]=="0"):
        column12="I"
        #call detect to take photos
        detect()
        #call identify with the path and the name of the column
        identify(path,column12)

    if(hello12[4:5]=="2"):
        column12="II"
        #call detect to take photos
        detect()
        #call identify with the path and the name of the column
        identify(path,column12)

    if(hello12[4:5]=="4"):
        column12="III"
        #call detect to take photos
        detect()
        #call identify with the path and the name of the column
        identify(path,column12)

    if(hello12[4:5]=="6"):
        column12="IV"
        #call detect to take photos
        detect()
        #call identify with the path and the name of the column
        identify(path,column12)

    if(hello12[4:5]=="8"):
        column12="V"
        #call detect to take photos
        detect()
        #call identify with the path and the name of the column
        identify(path,column12)
